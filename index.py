# importer module de qt5
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.uic import loadUiType 
import sqlite3
import datetime
from openpyxl import Workbook  
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
#importer le module qui permmet de relier le fichier deseigner avec le code python
#loadUiType
import sys
import sqlite3
from sqlite3 import Error

#determiner le lien du fichier QTdesigner
MainUi,_ = loadUiType('main.ui')
## creation de la classe principale
class Main(QMainWindow , MainUi):
    def __init__(self, parent=None) :
        ##pour overriwte le deseigner
        super(Main, self).__init__(parent) 
        QMainWindow.__init__(self)
        self.setupUi(self)
        self.loginid.setVisible(False)
        self.Db_Connect()
        self.Handel_boutons()
        self.UI_Changes()
        self.menu1.setCurrentIndex(4)
        
####### traiter les modification des fenetre et boutons #####""               
    def UI_Changes(self):
            ##pour traiter les pages
        self.menu1.tabBar().setVisible(False)
            # champ cheque et carte  ajout recu
        self.label_21.setVisible(False)
        self.lineEdit_17.setVisible(False)
        self.label_22.setVisible(False)
        self.lineEdit_18.setVisible(False)
        self.label_23.setVisible(False)
        self.dateEdit.setVisible(False)
        self.label_11.setVisible(False)
        self.lineEdit_10.setVisible(False)
        self.label_27.setVisible(False)
        self.lineEdit_20.setVisible(False)
        ## atribution date du jour champ echeance ajout recu
        self.dateEdit.setDate(datetime.date.today()) 
        ########### masquer mot de passe ################
        self.mdp_text.setEchoMode(QLineEdit.Password)
        self.lineEdit_4.setEchoMode(QLineEdit.Password)
        self.lineEdit_5.setEchoMode(QLineEdit.Password)
    
    def ouvrir_menu_utilisateur(self):
        self.menu1.setCurrentIndex(0)
        self.menu1_1.setCurrentIndex(0)
    
    def ouvrir_menu_recu(self):
        self.menu1.setCurrentIndex(1)
        self.menu1_2.setCurrentIndex(0)
        
    def ouvrir_menu_caisse(self):
        self.menu1.setCurrentIndex(2)
        self.menu1_3.setCurrentIndex(0)
    
    def ouvrir_menu_rapport(self):
        self.menu1.setCurrentIndex(3)

    def ouvrir_menu_fermer(self):
        self.menu1.setCurrentIndex(1)    

  ########################connection avec la base de données######################################################""
    def Db_Connect(self):
            self.db = sqlite3.connect('chifacaisse.db')
            self.cur = self.db.cursor()            
  ############################################ traitement des actions des boutons ################################"
    def Handel_boutons(self):
        self.recus_btn.clicked.connect(self.ouvrir_menu_recu)
        self.utilisateur_btn.clicked.connect(self.ouvrir_menu_utilisateur)
        self.caisse_btn.clicked.connect(self.ouvrir_menu_caisse)
        self.rapport_btn.clicked.connect(self.ouvrir_menu_rapport)
        self.valider_mdp.clicked.connect(self.Valider_Mot_de_passe)
        self.fermer_btn.clicked.connect(self.fermetur_app)
        self.fermer_app1.clicked.connect(self.fermetur_app)
        self.recherche_utilisateur.clicked.connect(self.recherche_users)
        self.modifier_utilisateur_btn.clicked.connect(self.modifier_utilisateur)
        self.ajouter_utilisateur_btn.clicked.connect(self.Ajout_utilisateur)
        self.controle_admin.stateChanged.connect(self.cocher_tt_permission2)
        self.checkBox_10.stateChanged.connect(self.cocher_tt_permission1)
        self.pushButton_33.clicked.connect(self.valider_permission)
        self.pushButton_11.clicked.connect(self.ajout_des_recus)
        self.comboBox_3.currentIndexChanged.connect(self.mode_de_reglement)
        self.comboBox_4.currentIndexChanged.connect(self.mode_de_reglement2)
        self.pushButton_16.clicked.connect(self.recherche_recu)
        self.imprimer_recu.clicked.connect(self.impression_des_recus_recherche)
        self.modifier_recu.clicked.connect(self.modifier_recus)
        self.supprimer_recu.clicked.connect(self.supprimer_les_recus)        
        #self.menu1_2.currentChanged(2).connect(self.utilisateur_liste)
        self.recherche_btn3.clicked.connect(self.recherche_liste_recus)

 ############## traitement du login ################################################################################           
    def Valider_Mot_de_passe(self):
        utilisateur = self.utilisateur_text.text()
        mdp =  self.mdp_text.text()
        self.cur.execute('''SELECT * FROM utilisateur WHERE nom=?''',(utilisateur,))
        data = self.cur.fetchone()
        if not data :
        ##" si pas d'enregistrement trouver"
         QMessageBox.warning(self,'LOGIN',"Utilisateur ou mot de passe Incorrect")
         self.utilisateur_text.setText("")
         self.mdp_text.setText("")        
        else :
            self.menu_index.setTitle("Session ouvert par :  " + data[1])
            self.loginid.setText(str(data[0]))
            ### si un enregistrement trouver
            ### recuperer le id de la base de donnée
            id = data[0]
            self.cur.execute('''SELECT * FROM permission WHERE utilisateur_id=?''', (id,))
            result = self.cur.fetchone()
            self.menu_index.setEnabled(True)
            self.menu1.setCurrentIndex(1)
            self.menu1_2.setCurrentIndex(0)
            self.utilisateur_text.setText("")
            self.mdp_text.setText("")
            ## permission utilisateur
            if result[2] == True :
                self.utilisateur_btn.setEnabled(True)
            else:
                self.utilisateur_btn.setEnabled(False)
            #### permission recu
            count_recu =0 
            if result[3] == True :
                self.recus_btn.setEnabled(True)
                count_recu += 1
            else: 
                self.recus_btn.setEnabled(False)
            if result[4] == True :
                self.modifier_recu.setEnabled(True)
                count_recu +=1
            else :
                self.modifier_recu.setEnabled(False)
                    
            if result[5] == True :
                self.supprimer_recu.setEnabled(True)
                count_recu +=1
            else :
                self.supprimer_recu.setEnabled(False)
                    
            if result[6] == True :
                self.imprimer_recu.setEnabled(True)
                self.imprimer_recu1.setEnabled(True)
                count_recu +=1
            else :
                self.imprimer_recu.setEnabled(False)
                self.imprimer_recu1.setEnabled(False)
                        
            if result[7] == True :
                self.exporter_recu.setEnabled(True)
                count_recu +=1
            else :
                self.exporter_recu.setEnabled(False)              
            if count_recu == 0 :
                self.recus_btn.setEnabled(False)
                
            ### permission caisse
            caisse_coun=0
            if result[8] == True :
                self.cloturer_caisse.setEnabled(True)
                caisse_coun+=1
            else:
                self.cloturer_caisse.setEnabled(False)
                    
            if result[9] == True :
                self.ouvrir_caisse.setEnabled(True)
                caisse_coun+=1
            else:
                self.ouvrir_caisse.setEnabled(False)
                    
            if result[10] == True :
                self.imprimer_caisse.setEnabled(True)
                self.imprimer_caisse1.setEnabled(True)
                self.imprimer_caisse2.setEnabled(True)
                caisse_coun+=1
            else:
                self.imprimer_caisse.setEnabled(False)
                self.imprimer_caisse1.setEnabled(False)
                self.imprimer_caisse2.setEnabled(False)
                        
            if result[11] == True :
                self.exporter_caisse.setEnabled(True)
                self.exporter_caisse1.setEnabled(True)
                self.exporter_caisse2.setEnabled(True)
                caisse_coun+=1
            else :
                self.exporter_caisse.setEnabled(False)
                self.exporter_caisse1.setEnabled(False)
                self.exporter_caisse2.setEnabled(False)
                    
            if caisse_coun == 0:
                self.caisse_btn.setEnabled(False) 
                
            ### permission rapport
            rapport_coun=0
            if result[12] == True :
                self.rapport_recherche.setEnabled(True)
                rapport_coun+=1
            else:
                self.rapport_recherche.setEnabled(False)
            
            if result[13] == True :
                self.imprimer_rapport.setEnabled(True)
                caisse_coun+=1
            else:
                self.imprimer_rapport.setEnabled(False)
                    
            if result[14] == True :
                self.envoie_rapport.setEnabled(True)
                caisse_coun+=1
            else:
                self.envoie_rapport.setEnabled(False)
                        
            if result[15] == True :
                self.exporte_rapport.setEnabled(True)
                caisse_coun+=1
            else :
                self.exporte_rapport.setEnabled(False)
    
            if rapport_coun == 0:
                self.caisse_btn.setEnabled(False)   
    ################################################ Utilisateur ###################################################
    def Ajout_utilisateur(self):
        erreur = 0
        if self.lineEdit_2.text() == "" :
            erreur += 1
        if self.lineEdit_4.text() == "" :
            erreur +=1
        if self.lineEdit_8.text() == "" :
            erreur += 1
            
        if erreur == 0 :   
            if self.lineEdit_5.text() == self.lineEdit_4.text() : 
                nom = self.lineEdit_2.text()
                mdp = self.lineEdit_4.text()
                mail = self.lineEdit_8.text()
                type_comp = self.comboBox.currentText()
                ### recherche si user existe deja dans la base
                self.cur.execute('''SELECT * FROM utilisateur WHERE nom=?''',(nom,))
                result = self.cur.fetchone()
                if not result  :
                    #si resulta de la recherche vide
                    valeur = (nom, mail, type_comp,mdp)
                    sql = """INSERT INTO utilisateur (nom, mail, compte_type, mot_de_passe) VALUES (?,?,?,?) """
                    self.cur.execute(sql, valeur)
                    self.db.commit()
                    ## rechercche le dernier utili ajouter
                    self.cur.execute('''SELECT * FROM utilisateur WHERE nom=?''',(nom,))
                    result = self.cur.fetchone()
                    #id_ut = result[0]
                    if type_comp == "admin" :
                        #utilisateur_id = id_ut
                        ajout_utilisateur= True 
                        ajout_recu= True
                        modification_recu= True
                        supprimer_recu= True
                        imprimer_recu= True
                        exporter_recu= True
                        cloturer_caisse= True
                        ouvrir_caisse= True
                        imprimer_caisse= True
                        exporter_caisse= True
                        recherche_rapport= True
                        imprimer_rapport= True
                        envoie_rapport= True
                        exporter_rapport= True
                        valeur = (result[0], ajout_utilisateur, 
                                ajout_recu,modification_recu,supprimer_recu,imprimer_recu,exporter_recu,
                                cloturer_caisse,ouvrir_caisse,imprimer_caisse,exporter_caisse,
                                recherche_rapport,imprimer_rapport,envoie_rapport,exporter_rapport)
                        sql = """INSERT INTO permission (utilisateur_id, ajout_utilisateur, 
                        ajout_recu, modification_recu, supprimer_recu, imprimer_recu, exporter_recu,	
                        cloturer_caisse, ouvrir_caisse, imprimer_caisse, exporter_caisse, recherche_rapport,
                        imprimer_rapport, envoie_rapport, exporter_rapport ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) """
                        self.cur.execute(sql, valeur)
                        self.db.commit()
                        QMessageBox.information(self,'Nouveau','Utilisateur Ajouter avec Succe')
                        self.lineEdit_2.setText("")
                        self.lineEdit_4.setText("")
                        self.lineEdit_5.setText("")
                        self.lineEdit_8.setText("")
                    else :
                        self.menu1_1.setCurrentIndex(2)
                        self.lineEdit_30.setText(result[1])
                        self.lineEdit_2.setText("")
                        self.lineEdit_4.setText("")
                        self.lineEdit_5.setText("")
                        self.lineEdit_8.setText("")
                else :
                    QMessageBox.warning(self,'Nouveau','utilisateur existe Deja') 
                    self.lineEdit_2.setText("")
                    self.lineEdit_4.setText("")
                    self.lineEdit_8.setText("")
            else : 
                QMessageBox.warning(self,"MOT DE PASSE", "Merci de VAlider Votre MOT DE PASSE")
                self.lineEdit_5.setText("")
        else :
            QMessageBox.warning(self,'Nouveau','Veuillez remplir tous les champs')  

    
    def valider_permission(self):
        nom = self.lineEdit_30.text()
        self.cur.execute('''SELECT * FROM utilisateur WHERE nom=?''',(nom,))
        result = self.cur.fetchone()
        if not result :
            QMessageBox.warning(self,'Ajout','veuillez ajouter un nouveau Utilisateur')
            self.menu1_1.setCurrentIndex(0)
        else :
            id_per = result[0]
            ut_ajout = self.checkBox.isChecked()
                    #### permission recu
            rec_ajout= self.checkBox_4.isChecked()
            rec_modification =self.checkBox_5.isChecked()
            rec_supprimer = self.checkBox_6.isChecked()
            rec_imp = self.checkBox_7.isChecked()
            rec_exp = self.checkBox_16.isChecked()
                    ### permission caisse
            cai_cloture = self.checkBox_8.isChecked()
            cai_ouvrir = self.checkBox_9.isChecked()
            cai_imp=self.checkBox_11.isChecked()
            cai_exp = self.checkBox_17.isChecked()
            rap_rech = self.rech_rapport_2.isChecked()
            rap_imp= self.checkBox_12.isChecked()
            rap_env = self.checkBox_13.isChecked()
            rap_exp = self.checkBox_18.isChecked()
            admin_check = self.controle_admin.isChecked()
            valeur_perm = (id_per, ut_ajout, 
                            rec_ajout, rec_modification, rec_supprimer, rec_imp, rec_exp,
                            cai_cloture, cai_ouvrir, cai_imp, cai_exp,
                            rap_rech, rap_imp, rap_env, rap_exp)
            sql = """INSERT INTO permission (utilisateur_id, ajout_utilisateur, 
                        ajout_recu, modification_recu, supprimer_recu, imprimer_recu, exporter_recu,	
                        cloturer_caisse, ouvrir_caisse, imprimer_caisse, exporter_caisse, recherche_rapport,
                        imprimer_rapport, envoie_rapport, exporter_rapport ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) """
            self.cur.execute(sql, valeur_perm)
            self.db.commit()   
            QMessageBox.information(self,'Nouveau','Permission ajouter avec Succe') 
            self.lineEdit_30.setText("")
            self.checkBox.setChecked(False)
            self.checkBox_4.setChecked(False)
            self.checkBox_5.setChecked(False)
            self.checkBox_6.setChecked(False)
            self.checkBox_7.setChecked(False)
            self.checkBox_16.setChecked(False)
            self.checkBox_8.setChecked(False)
            self.checkBox_9.setChecked(False)
            self.checkBox_11.setChecked(False)
            self.checkBox_17.setChecked(False)
            self.checkBox_12.setChecked(False)
            self.checkBox_13.setChecked(False)
            self.checkBox_18.setChecked(False)
            self.rech_rapport_2.setChecked(False)
            self.menu1_1.setCurrentIndex(0)    

    def cocher_tt_permission1(self):
        if self.checkBox_10.isChecked() == True:
            self.checkBox.setChecked(True)
            self.checkBox_4.setChecked(True)
            self.checkBox_5.setChecked(True)
            self.checkBox_6.setChecked(True)
            self.checkBox_7.setChecked(True)
            self.checkBox_16.setChecked(True)
            self.checkBox_8.setChecked(True)
            self.checkBox_9.setChecked(True)
            self.checkBox_11.setChecked(True)
            self.checkBox_17.setChecked(True)
            self.checkBox_12.setChecked(True)
            self.checkBox_13.setChecked(True)
            self.checkBox_18.setChecked(True)
            self.rech_rapport_2.setChecked(True)
        else :
            self.checkBox.setChecked(False)
            self.checkBox_4.setChecked(False)
            self.checkBox_5.setChecked(False)
            self.checkBox_6.setChecked(False)
            self.checkBox_7.setChecked(False)
            self.checkBox_16.setChecked (False)
            self.checkBox_8.setChecked(False)
            self.checkBox_9.setChecked(False)
            self.checkBox_11.setChecked(False)
            self.checkBox_17.setChecked(False)
            self.checkBox_12.setChecked(False)
            self.checkBox_13.setChecked(False)
            self.checkBox_18.setChecked(False)
            self.rech_rapport_2.setChecked(False)
            
    def cocher_tt_permission2(self):
        if self.controle_admin.isChecked() == True:
            self.aj_utilisateur.setChecked(True)
            self.mod_utilisateur.setChecked(True)
            self.sup_utilisateur.setChecked(True)
            self.aj_recu.setChecked(True)
            self.mod_recu.setChecked(True)
            self.sup_recu.setChecked(True)
            self.imp_recu.setChecked(True)
            self.exp_recu.setChecked(True)
            self.clo_caisse.setChecked(True)
            self.ouv_caisse.setChecked(True)
            self.imp_caisse.setChecked(True)
            self.exp_caisse.setChecked(True)
            self.rech_rapport.setChecked(True)
            self.imp_rapport.setChecked(True)
            self.env_rapport.setChecked(True)
            self.exp_rapport.setChecked(True)
        else :
            self.aj_utilisateur.setChecked(False)
            self.mod_utilisateur.setChecked(False)
            self.sup_utilisateur.setChecked(False)
            self.aj_recu.setChecked(False)
            self.mod_recu.setChecked(False)
            self.sup_recu.setChecked(False)
            self.imp_recu.setChecked(False)
            self.exp_recu.setChecked(False)
            self.clo_caisse.setChecked(False)
            self.ouv_caisse.setChecked(False)
            self.imp_caisse.setChecked(False)
            self.exp_caisse.setChecked(False)
            self.rech_rapport.setChecked(False)
            self.imp_rapport.setChecked(False)
            self.env_rapport.setChecked(False)
            self.exp_rapport.setChecked(False)
            
     
    def recherche_users(self):
        recherche = self.utilisateur_text_2.text()
        self.cur.execute('''SELECT * FROM utilisateur WHERE nom=?''',(recherche,))
        result = self.cur.fetchone()
        if not result :
            QMessageBox.warning(self,'Recherche',"Utilisateur non trouvé")
        else:
            self.util_modification.setText(result [1])
            self.mdp_utilisateur.setText(result [4])
            self.mail_utilisateur.setText(result [2])
            ##### recherche permission
            ## permission utilisateur
            id = result[0]
            self.cur.execute('''SELECT * FROM permission WHERE utilisateur_id=?''', (id,))
            result = self.cur.fetchone()
            adminper=0
            if result[2] == True :
                self.aj_utilisateur.setChecked(True)
                self.mod_utilisateur.setChecked(True)
                self.sup_utilisateur.setChecked(True)
            else:
                self.aj_utilisateur.setChecked(False)
                self.mod_utilisateur.setChecked(False)
                self.sup_utilisateur.setChecked(False)
                adminper+=1
            #### permission recu
            count_recu =0 
            if result[3] == True :
                self.aj_recu.setChecked(True)
            else: 
                self.aj_recu.setChecked(False)
                adminper +=1
            if result[4] == True :
                self.mod_recu.setChecked(True)
            else :
                self.mod_recu.setChecked(False)
                adminper+=1
            if result[5] == True :
                self.sup_recu.setChecked(True)
            else :
                self.sup_recu.setChecked(False)
                adminper +=1    

            if result[6] == True :
                self.imp_recu.setChecked(True)

            else :
                self.imp_recu.setChecked(False)
                adminper += 1
                        
            if result[7] == True :
                self.exp_recu.setChecked(True)
            else :
                self.exp_recu.setChecked(False)              
                adminper += 1
                           
            ### permission caisse
            if result[8] == True :
                self.clo_caisse.setChecked(True)
            else:
                self.clo_caisse.setChecked(False)
                adminper += 1
            if result[9] == True :
                self.ouv_caisse.setChecked(True)
            else:
                self.ouv_caisse.setChecked(False)
                adminper += 1
            if result[10] == True :
                self.imp_caisse.setChecked(True)
            else:
                self.imp_caisse.setChecked(False)
                adminper += 1
            if result[11] == True :
                self.exp_caisse.setChecked(True)
            else :
                self.exp_caisse.setChecked(False)
                adminper += 1
            ### permission rapport
            if result[12] == True :
                self.rech_rapport.setChecked(True)
            else:
                self.rech_rapport.setChecked(False)
                adminper +=1            
            if result[13] == True :
                self.imp_rapport.setChecked(True)
            else:
                self.imp_rapport.setChecked(False)
                adminper += 1
            if result[14] == True :
                self.env_rapport.setChecked(True)
            else:
                self.env_rapport.setChecked(False)
                adminper += 1
                        
            if result[15] == True :
                self.exp_rapport.setChecked(True)
            else :
                self.exp_rapport.setChecked(False)
                adminper +=1
                            
            if adminper == 0 :
                self.controle_admin.setChecked(True)
            else :
                self.controle_admin.setChecked(False)
            
    def modifier_utilisateur(self) :
        recherche = self.utilisateur_text_2.text()
        self.cur.execute('''SELECT * FROM utilisateur WHERE nom=?''',(recherche,))
        result = self.cur.fetchone()
        if not(result) == False :
            id_ut = result[0]
            type_comp = result[3]
            nom = self.util_modification.text()
            mdp = self.mdp_utilisateur.text()
            mail = self.mail_utilisateur.text()
            valeur_util = (nom, mail, type_comp, mdp, id_ut)
            ##### recherche permission
            ## permission utilisateur
            self.cur.execute('''SELECT * FROM permission WHERE utilisateur_id=?''', (id_ut,))
            result = self.cur.fetchone()
            id_per = result[0]
            ut_ajout = self.aj_utilisateur.isChecked()
            ut_modification =self.mod_utilisateur.isChecked()
            ut_supprimer = self.sup_utilisateur.isChecked()
                #### permission recu
            rec_ajout= self.aj_recu.isChecked()
            rec_modification =self.mod_recu.isChecked()
            rec_supprimer = self.sup_recu.isChecked()
            rec_imp = self.imp_recu.isChecked()
            rec_exp = self.exp_recu.isChecked()
                        ### permission caisse
            cai_cloture = self.clo_caisse.isChecked()
            cai_ouvrir = self.ouv_caisse.isChecked()
            cai_imp=self.imp_caisse.isChecked()
            cai_exp = self.exp_caisse.isChecked()
            rap_rech = self.rech_rapport.isChecked()
            rap_imp= self.imp_rapport.isChecked()
            rap_env = self.env_rapport.isChecked()
            rap_exp = self.exp_rapport.isChecked()
            admin_check = self.controle_admin.isChecked()
            valeur_perm = (ut_ajout, 
                        rec_ajout, rec_modification, rec_supprimer, rec_imp, rec_exp,
                        cai_cloture, cai_ouvrir, cai_imp, cai_exp,
                        rap_rech, rap_imp, rap_env, rap_exp, id_per)
            sql_ut = """UPDATE utilisateur SET nom =?, mail=?,compte_type=?, mot_de_passe =?  WHERE id = ?"""
            sql_per = """UPDATE permission SET ajout_utilisateur = ? ,
                        ajout_recu = ?, modification_recu = ?, supprimer_recu=?, imprimer_recu=?, exporter_recu=?,
                        cloturer_caisse=?, ouvrir_caisse=?, imprimer_caisse=?, exporter_caisse=?,
                        recherche_rapport=?, imprimer_rapport=?, envoie_rapport=?, exporter_rapport=? WHERE id = ? """
            self.cur.execute(sql_ut,valeur_util)
            self.cur.execute(sql_per,valeur_perm)
            self.db.commit()
            QMessageBox.information(self,'Modification','Utilisateur Modifier avec Succe')
        else :
            QMessageBox.information(self,'Recherche','Veuillez chercher un Utilisateur a Modifier')   
        
############################################## RECUS ###############################################################

    def mode_de_reglement(self):
        if self.comboBox_3.currentText() == "ESPECE" :
            self.label_21.setVisible(False)
            self.lineEdit_17.setVisible(False)
            self.label_22.setVisible(False)
            self.lineEdit_18.setVisible(False)
            self.label_23.setVisible(False)
            self.dateEdit.setVisible(False)
                    
        if self.comboBox_3.currentText() == "CHEQUE" :
            self.label_21.setVisible(True)
            self.lineEdit_17.setVisible(True)
            self.dateEdit.setVisible(True)
            self.label_23.setVisible(True)
            QMessageBox.information(self,"CHEQUE","Saisir le NUMERO DU CHEQUE")
        else :
            self.label_21.setVisible(False)
            self.lineEdit_17.setVisible(False)
            self.dateEdit.setVisible(False)
  
        if self.comboBox_3.currentText() == "CARTE" :
            self.label_22.setVisible(True)
            self.lineEdit_18.setVisible(True)
            QMessageBox.information(self,"CARTE","Saisir le NUMERO DE LA CARTE")
        else :
            self.label_22.setVisible(False)
            self.lineEdit_18.setVisible(False)
            
    def mode_de_reglement2(self):
        if self.comboBox_4.currentText() == "ESPECE" :
            self.label_36.setVisible(False)
            self.label_37.setVisible(False)
            self.label_39.setVisible(False)
            self.lineEdit_27.setVisible(False)
            self.lineEdit_28.setVisible(False)
            self.echeance.setVisible(False)

        if self.comboBox_4.currentText() == "CHEQUE" :
            self.label_36.setVisible(True)
            self.label_37.setVisible(False)
            self.label_39.setVisible(True)
            self.lineEdit_27.setVisible(True)
            self.lineEdit_28.setVisible(False)
            self.echeance.setVisible(True)
            QMessageBox.information(self,"CHEQUE","Saisir le NUMERO DU CHEQUE  +  ECHEANCE")
  
        if self.comboBox_4.currentText() == "CARTE" :
            self.label_36.setVisible(False)
            self.label_37.setVisible(True)
            self.label_39.setVisible(False)
            self.lineEdit_27.setVisible(False)
            self.lineEdit_28.setVisible(True)
            self.echeance.setVisible(False)
            QMessageBox.information(self,"CARTE","Saisir le NUMERO D'AUTORISATION")
    
    def ajout_des_recus (self):
        erreur = ""
        sql= self.cur.execute("SELECT * FROM recus ORDER BY num DESC LIMIT 1")
        resulta = self.cur.fetchone()
        if not resulta :
            num_recu = 100001
        else :
            num_recu = int(resulta[0]) + 1
        
        self.lineEdit_10.setText(str(num_recu))
        date_creation = datetime.date.today().isoformat()
        if self.lineEdit_11.text != "" :
            patient = self.lineEdit_11.text()
        else :
            erreur ="Nom du patient"
            
        if self.lineEdit_13.text()!= "" :
                az_num = self.lineEdit_13.text()
        else :
            erreur = erreur +  " + "  + " Numero AZ"
            
        if self.lineEdit_12.text() != "" :
            organisme = self.lineEdit_12.text()
        else : 
            erreur = erreur +  " + "  + " Organizme"
            
        if self.lineEdit_15.text()!= "":
            motif = self.lineEdit_15.text()
        else :
            erreur = erreur +  " + "  + " Motif"
            
        if self.lineEdit_14.text()!="":
            medecin = self.lineEdit_14.text()
        else :
            erreur = erreur +  " + "  + "Medecin"
            
        if self.lineEdit_16.text() != "" :
            montant = self.lineEdit_16.text()
        else :
            erreur = erreur +  " + "  + " Montant"
        
        try :
            montant = int(montant)
        except:
            erreur = erreur + " + " + " Montant INVALIDE"

        echeance = self.dateEdit.date().toString("yyyy-MM-dd")
        reglement = self.comboBox_3.currentText()
        num_cheque = self.lineEdit_17.text()
        num_carte = self.lineEdit_18.text()
        observation = self.lineEdit_31.text()
        cloture = False
        ut_id= int(self.loginid.text())
        if erreur == "" :
            valeur =(num_recu,date_creation ,patient ,az_num ,organisme,motif,medecin ,montant ,
                 echeance,reglement,num_cheque ,num_carte,observation,cloture ,ut_id)
            sql = """INSERT INTO recus (num,date_creation,patient,Az_Numero,organisme_pec,
                motif,medecin,montant,echeance,mode_reglement,num_cheque,num_carte,
                observation,cloture,Utilisateur_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) """
            self.cur.execute(sql, valeur)
            self.db.commit()
            QMessageBox.information(self,"RECU NUMERO", "RECU NUMERO :" + str(num_recu) + " Ajouter avec Succee")
            self.label_11.setVisible(True)
            self.lineEdit_10.setVisible(True)
            imp_erreur = False
            while imp_erreur == False :
                try :
                    doc = load_workbook('recus.xlsx')
                    sheet = doc.active
                    img = Image("icons/chifa_hor.png")
                    img_ok = sheet._images
                    if len(img_ok) == 0 :
                        sheet.add_image(img , 'B1')    
                    sheet['C6'] = self.lineEdit_10.text()
                    sheet['F6'] = datetime.date.today().isoformat()
                    sheet['C8'] = self.lineEdit_11.text()
                    sheet['C10'] = self.lineEdit_12.text()
                    sheet['C12'] = self.lineEdit_15.text()
                    sheet['C14'] = self.lineEdit_16.text() + "  DH"
                    sheet['C16'] = self.comboBox_3.currentText()
                    if self.comboBox_3.currentText()== "CHEQUE" :
                        sheet['F16'] = self.lineEdit_17.text() + " / " + self.dateEdit.date().toString("yyyy-MM-dd")
                    elif self.comboBox_3.currentText()== "CARTE":        
                        sheet['F17'] = self.lineEdit_18.text()
                    else : 
                        sheet['F16'] = ""
                        sheet['F17'] = ""
                    ############
                    doc.save('recus.xlsx')
                    os.startfile('recus.xlsx', 'print')
                    imp_erreur = True
                except :
                    QMessageBox.warning(self,"ERREUR","Erreur d'impression")

            ##########
            self.lineEdit_10.setText("")
            self.lineEdit_10.setVisible(False)
            self.label_11.setVisible(False)
            self.lineEdit_11.setText("")
            self.lineEdit_13.setText("")
            self.lineEdit_12.setText("")
            self.lineEdit_15.setText("")
            self.lineEdit_14.setText("")
            self.lineEdit_16.setText("")
            self.dateEdit.setDate(datetime.date.today()) 
            self.comboBox_3.setCurrentIndex(0)
            self.lineEdit_17.setText("")
            self.lineEdit_18.setText("")
            self.lineEdit_31.setText("")
        else :
            QMessageBox.warning(self,"Recu", "Veuillez Verifier ces Champs  : " + erreur )

    def impression_des_recus_recherche (self):
        imp_erreur = False
        while imp_erreur ==False :
            try :                
                doc = load_workbook('recus.xlsx')
                sheet = doc.active
                img = Image("icons/chifa_hor.png")
                img_ok = sheet._images
                print (len(img_ok))
                if len(img_ok) == 0 :
                    sheet.add_image(img , 'B1')
                    
                sheet['C6'] = self.lineEdit_19.text()
                sheet['F6'] = self.dateEdit_3.date().toString("yyyy-MM-dd")
                sheet['C8'] = self.lineEdit_21.text()
                sheet['C10'] = self.lineEdit_22.text()
                sheet['C12'] = self.lineEdit_25.text()
                sheet['C14'] = self.lineEdit_26.text() + "  DH"
                sheet['C16'] = self.comboBox_4.currentText()
                if self.comboBox_4.currentText()== "CHEQUE" :
                    sheet['F16'] = self.lineEdit_27.text() + " / " + self.echeance.date().toString("yyyy-MM-dd")
                elif self.comboBox_4.currentText()== "CARTE" :
                    sheet['F17'] = self.lineEdit_28.text()
                else : 
                    sheet['F16'] = ""
                    sheet['F17'] = ""
                ###################################
                doc.save('recus.xlsx')
                os.startfile('recus.xlsx', 'print')
                ####################################
            except :
                QMessageBox.warning(self,"ERREUR","Erreur d'impression")

    def recherche_recu(self):
        num_recu = int(self.lineEdit_19.text())
        self.cur.execute('''SELECT * FROM recus WHERE num=?''',(num_recu,))
        resulta = self.cur.fetchone()
        if not resulta :
            QMessageBox.warning(self,"RECHERCHE","Num de RECU non TROUVER")
        else :
            self.groupBox_11.setEnabled(True)
            self.lineEdit_20.setText(str(resulta[14]))
            ##convertir str en date
            date_time = datetime.datetime.strptime(resulta[1], '%Y-%m-%d')
            self.dateEdit_3.setDate(date_time)
            self.lineEdit_21.setText(resulta[2])
            self.lineEdit_23.setText(resulta[3])
            self.lineEdit_22.setText(resulta[4])
            self.lineEdit_25.setText(resulta[5])
            self.lineEdit_24.setText(resulta[6])
            self.lineEdit_26.setText(str(resulta[7]))
            ##### mode de reglement ##############
            if resulta[9] == "ESPECE" :
                self.comboBox_4.setCurrentIndex(0)
                self.label_36.setVisible(False)
                self.label_37.setVisible(False)
                self.label_39.setVisible(False)
                self.lineEdit_27.setVisible(False)
                self.lineEdit_28.setVisible(False)
                self.echeance.setVisible(False)
            elif resulta[9] == "CHEQUE" :
                self.comboBox_4.setCurrentIndex(1)
                self.label_36.setVisible(True)
                self.label_39.setVisible(True)
                self.lineEdit_27.setVisible(True)
                self.echeance.setVisible(True)
                self.label_37.setVisible(False)
                self.lineEdit_28.setVisible(False)
        
            else :
                self.comboBox_4.setCurrentIndex(2)
                self.label_37.setVisible(True)
                self.lineEdit_28.setVisible(True)
                self.label_36.setVisible(False)
                self.label_39.setVisible(False)
                self.lineEdit_27.setVisible(False)
                self.echeance.setVisible(False)
            #######################################    
            date_time = datetime.datetime.strptime(resulta[8], '%Y-%m-%d')
            self.echeance.setDate(date_time)
            self.lineEdit_27.setText(resulta[10])
            self.lineEdit_28.setText(resulta[11])
            self.lineEdit_32.setText(resulta[12])
            if resulta[13] == True :
                self.modifier_recu.setVisible(False)
                self.supprimer_recu.setVisible(False)
                self.label_27.setVisible(True)
                self.groupBox_11.setEnabled(False)

    def modifier_recus (self):
        if self.lineEdit_20.text() == self.loginid.text():
            num_recu = int(self.lineEdit_19.text())
            date_creation = self.dateEdit_3.date().toString("yyyy-MM-dd")
            erreur = ""
            if self.lineEdit_21.text != "" :
                patient = self.lineEdit_21.text()
            else :
                erreur ="Nom du patient"
                
            if self.lineEdit_23.text()!= "" :
                    az_num = self.lineEdit_23.text()
            else :
                erreur = erreur +  " + "  + " Numero AZ"
                
            if self.lineEdit_22.text() != "" :
                organisme = self.lineEdit_22.text()
            else : 
                erreur = erreur +  " + "  + " Organizme"
                
            if self.lineEdit_25.text()!= "":
                motif = self.lineEdit_25.text()
            else :
                erreur = erreur +  " + "  + " Motif"
                
            if self.lineEdit_24.text()!="":
                medecin = self.lineEdit_24.text()
            else :
                erreur = erreur +  " + "  + "Medecin"
                
            if self.lineEdit_26.text() != "" :
                montant = self.lineEdit_26.text()
            else :
                erreur = erreur +  " + "  + " Montant"
            
            try :
                montant = int(montant)
            except:
                self.lineEdit_26.setText("")
                erreur = erreur + " + " + " Montant INVALIDE"

            echeance = self.echeance.date().toString("yyyy-MM-dd")
            reglement = self.comboBox_4.currentText()
            num_cheque = self.lineEdit_27.text()
            num_carte = self.lineEdit_28.text()
            observation = self.lineEdit_32.text()
            cloture = False
            ut_id= int(self.loginid.text())
            if erreur == "" :
                valeur =(num_recu, date_creation ,patient ,az_num ,organisme,motif,medecin ,montant ,
                    echeance,reglement,num_cheque ,num_carte,observation,cloture, ut_id,num_recu)
                print(valeur)
                sql = """UPDATE recus SET num=?, date_creation=?,patient = ?,Az_Numero=? ,organisme_pec=?,
                    motif=?,medecin=?,montant=?,echeance=?,mode_reglement=?,num_cheque=?,num_carte=?,
                    observation=?,cloture=?, Utilisateur_id =? WHERE num = ?"""
                self.cur.execute(sql, valeur)
                self.db.commit()
                QMessageBox.information(self,"MODIFICATION", "RECU NUMERO :" + str(num_recu) + " MODIFIE avec Succès")
            else : 
                QMessageBox.warning(self,"ERREUR", "Veuillez Verifier ces Champs  : " + erreur )
        else :
            QMessageBox.warning(self,"ERREUR","Vous Pouvez pas Modifier Ce Recu Probleme D'Identification")

    def supprimer_les_recus(self):
        num_recu = int(self.lineEdit_19.text())
        self.cur.execute('DELETE FROM recus WHERE num=?', (num_recu,))
        self.db.commit()
        QMessageBox.information(self,"Suppression","RECU Bien Supprimer")
        self.lineEdit_19.setText("")
        self.lineEdit_21.setText("")
        self.lineEdit_23.setText("")
        self.lineEdit_22.setText("")
        self.lineEdit_25.setText("")
        self.lineEdit_24.setText("")
        self.lineEdit_26.setText("")
        self.dateEdit.setDate(datetime.date.today()) 
        self.comboBox_4.setCurrentIndex(0)
        self.lineEdit_27.setText("")
        self.lineEdit_28.setText("")
        self.lineEdit_32.setText("")
    
    def recherche_liste_recus (self):
        pass


########################################################################################
    def fermetur_app(self):
        QApplication.quit()
          
##########################################################
# cretaion d'une fonction pour recuperer qydesigner
def main():
    # app pour recuperer objet Qapplication pour traiter avec le terminal
    app = QApplication(sys.argv) 
    # recuper un objet de la classe principale
    fenetre = Main()
    fenetre.show() #pour afficher la fenetre
    app.exec_() # la boucle infini pour que la fenetre resste afficher
    
# creation condition pour que le programme appele la fonction 
# de lancement main()
if __name__ == '__main__' :
    main()